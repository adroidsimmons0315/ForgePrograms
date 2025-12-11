import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from datetime import date
from copy import copy

import pandas as pd
from openpyxl import load_workbook

import shared_functions as sf

STYLE_TEMPLATE_ROW = 2  # first data row under headers


def copy_cell_styles(ws, src_row: int, dst_row: int, cols: list[str]) -> None:
    """
    Copy the entire style object from src_row to dst_row for the given columns.
    Using dst._style = copy(src._style) avoids StyleProxy / hashability issues.
    """
    for col in cols:
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        dst._style = copy(src._style)


def sort_worksheet_by_column_a(ws) -> None:
    """
    Sort the data rows (row 2 .. max_row) by column A (ascending).
    """
    max_row = ws.max_row
    max_col = ws.max_column

    data_rows = list(
        ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True)
    )

    def key_func(row):
        val = row[0]  # column A
        try:
            return int(val)
        except (TypeError, ValueError):
            return float("inf")

    data_rows.sort(key=key_func)

    for r_idx, row_values in enumerate(data_rows, start=2):
        for c_idx, value in enumerate(row_values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def normalize_shr_location(raw: str) -> str:
    """
    Ensure the location string starts with 'SHR: '.
    If the user types '1234', becomes 'SHR: 1234'.
    If they type 'SHR: 1234', kept as is.
    Case-insensitive on the prefix.
    """
    s = raw.strip()
    if not s:
        return s
    if s.upper().startswith("SHR:"):
        return s  # assume user knows what they're doing
    return f"SHR: {s}"


def run_scan_out(root: tk.Misc | None = None):
    """
    Scan items OUT of inventory.

    If found: update Z (Last Scanned), AA (Last Verified), X (In/Out) = "Out", and I (Loc).
    If not found or duplicate: append a new row with seq (50xxx/51xxx),
    Asset Id in column C, I = SHR-location, X = "Out", Z/AA = today.
    """
    base_dir = Path(__file__).resolve().parent

    # Ensure root exists so dialogs/windows have a parent
    owns_root = False
    if root is None:
        root = tk.Tk()
        root.withdraw()
        owns_root = True

    # Select Excel file
    data_path = sf.select_inventory_excel_file(root=root, title="Select Inventory Excel File (Scan OUT)")
    if not data_path:
        if owns_root:
            root.destroy()
        return

    data_path = Path(data_path)

    # Load data with reusable function (used only to locate rows)
    asset_column = "Asset Id"
    try:
        df_inventory = sf.load_inventory_dataframe(data_path, asset_column=asset_column)
    except Exception as e:
        messagebox.showerror("Load error", f"Failed to read Excel file:\n{e}", parent=root)
        if owns_root:
            root.destroy()
        return

    # Normalize identifiers in inventory (digits-only, pad to 6)
    df_inventory = df_inventory.copy()
    df_inventory["Normalized"] = df_inventory[asset_column].map(sf.normalize)
    normalized_asset_set = set(df_inventory["Normalized"])

    # Open the workbook with openpyxl so we can edit specific cells / append rows
    try:
        wb = load_workbook(data_path)
    except Exception as e:
        messagebox.showerror("Load error", f"Failed to open workbook:\n{e}", parent=root)
        if owns_root:
            root.destroy()
        return

    ws = wb[wb.sheetnames[0]]  # assume first sheet is the inventory

    # Map from normalized code -> sheet row number (1-based Excel row index)
    normalized_to_row: dict[str, int] = {}
    for idx, code in df_inventory["Normalized"].items():
        if pd.isna(code):
            continue
        normalized_to_row[str(code)] = int(idx) + 2  # header is row 1

    # Build list of existing SHR locations from column I (only ones starting with "SHR:")
    loc_values: set[str] = set()
    for r in range(2, ws.max_row + 1):
        val = ws[f"I{r}"].value
        if isinstance(val, str):
            v = val.strip()
            if v.upper().startswith("SHR:"):
                loc_values.add(v)

    existing_locations = sorted(loc_values)

    # Determine starting sequence numbers for 50xxx (new items) and 51xxx (duplicates)
    next_new_seq = 50000
    next_dup_seq = 51000
    for r in range(2, ws.max_row + 1):
        val = ws[f"A{r}"].value
        try:
            n = int(str(val))
        except (TypeError, ValueError):
            continue
        if 50000 <= n < 51000 and n >= next_new_seq:
            next_new_seq = n + 1
        elif 51000 <= n < 52000 and n >= next_dup_seq:
            next_dup_seq = n + 1

    # Use a real date object so Excel applies its date format
    today = date.today()

    scanned: set[str] = set()
    new_items_count = 0
    duplicate_count = 0

    # UI window
    scan_win = tk.Toplevel(root)
    scan_win.title("Scan OUT")
    scan_win.geometry("720x600")
    scan_win.minsize(640, 460)

    def close_window():
        if owns_root:
            root.destroy()
        else:
            scan_win.destroy()

    scan_win.protocol("WM_DELETE_WINDOW", lambda: close_window())

    container = ttk.Frame(scan_win, padding=12)
    container.pack(fill="both", expand=True)

    # Location selector row
    loc_row = ttk.Frame(container)
    loc_row.pack(fill="x", pady=(0, 8))
    ttk.Label(loc_row, text="SHR Location:", font=("Segoe UI", 10, "bold")).pack(
        side="left", padx=(0, 8)
    )

    loc_var = tk.StringVar()
    loc_combo = ttk.Combobox(
        loc_row,
        textvariable=loc_var,
        values=existing_locations,
        state="normal",  # allow typing as well as dropdown
        width=40,
    )
    loc_combo.pack(side="left", fill="x", expand=True)
    if existing_locations:
        loc_combo.set(existing_locations[0])

    ttk.Label(
        container,
        text="Scan or type an asset code, then press Enter. "
             "Type 'done' to finish (Scan OUT).",
        font=("Segoe UI", 11),
        wraplength=680,
        justify="left",
    ).pack(anchor="w", pady=(0, 8))

    # Text area + scrollbar
    wrap = ttk.Frame(container)
    wrap.pack(fill="both", expand=True)
    yscroll = ttk.Scrollbar(wrap, orient="vertical")
    output_text = tk.Text(wrap, width=90, height=22, yscrollcommand=yscroll.set)
    yscroll.config(command=output_text.yview)
    output_text.pack(side="left", fill="both", expand=True)
    yscroll.pack(side="right", fill="y")

    def log(msg: str):
        output_text.insert(tk.END, msg + "\n")
        output_text.see(tk.END)

    # Entry row
    row = ttk.Frame(container)
    row.pack(fill="x", pady=8)
    ttk.Label(row, text="Code:").pack(side="left", padx=(0, 8))
    scan_entry = ttk.Entry(row, width=32, font=("Segoe UI", 14))
    scan_entry.pack(side="left")
    scan_entry.focus()

    # Buttons
    btns = ttk.Frame(container)
    btns.pack(fill="x", pady=(4, 0))

    def save_and_close():
        """Sort by column A, save workbook, and close the window."""
        sort_worksheet_by_column_a(ws)

        try:
            wb.save(data_path)
        except Exception as e:
            messagebox.showerror(
                "Save error",
                f"Failed to save changes to Excel file:\n{e}",
                parent=scan_win,
            )
            return

        loc_display = normalize_shr_location(loc_var.get().strip()) if loc_var.get().strip() else "(none)"
        log("\n--- Scan OUT Completed ---")
        log(f"SHR Location used: {loc_display}")
        log(f"Unique items scanned: {len(scanned)}")
        log(f"New items added (50xxx): {new_items_count}")
        log(f"Duplicate entries added (51xxx): {duplicate_count}")

        messagebox.showinfo(
            "Scan OUT updated",
            f"Changes saved to:\n{data_path}",
            parent=scan_win,
        )
        close_window()

    def handle_scan(event=None):
        nonlocal next_new_seq, next_dup_seq, new_items_count, duplicate_count

        code = scan_entry.get().strip()
        scan_entry.delete(0, tk.END)
        if not code:
            return

        raw_loc = loc_var.get().strip()
        if not raw_loc:
            messagebox.showwarning(
                "Location required",
                "Please select or type a SHR location before scanning.",
                parent=scan_win,
            )
            return

        loc = normalize_shr_location(raw_loc)

        # Keep the normalized form in the combobox / UI too
        loc_var.set(loc)

        if code.lower() == "done":
            save_and_close()
            return

        normalized_code = sf.normalize(code)

        if normalized_code in normalized_asset_set:
            # Item exists in inventory
            if normalized_code in scanned:
                # DUPLICATE: append at bottom with 51xxx sequence, X="Out", I=loc
                row_num = ws.max_row + 1
                seq = next_dup_seq
                next_dup_seq += 1
                duplicate_count += 1

                copy_cell_styles(ws, STYLE_TEMPLATE_ROW, row_num, ["A", "C", "I", "X", "Z", "AA"])

                ws[f"A{row_num}"].value = seq      # sequential 51xxx+
                ws[f"C{row_num}"].value = code     # Asset Id
                ws[f"I{row_num}"].value = loc
                ws[f"X{row_num}"].value = "Out"
                ws[f"Z{row_num}"].value = today
                ws[f"AA{row_num}"].value = today

                log(
                    f"[!] {code} duplicate ID. "
                    f"Added OUT row {row_num} with A={seq}, C={code}, I='{loc}', X=Out."
                )
            else:
                # FIRST SEEN in this session: update existing row
                scanned.add(normalized_code)
                row_num = normalized_to_row.get(normalized_code)
                if row_num is None:
                    log(f"[?] {code} found in inventory set, but row could not be located.")
                    return

                ws[f"Z{row_num}"].value = today
                ws[f"AA{row_num}"].value = today
                ws[f"X{row_num}"].value = "Out"
                ws[f"I{row_num}"].value = loc

                log(
                    f"[✓] {code} found in inventory. "
                    f"Set I{row_num}='{loc}', X{row_num}=Out, "
                    f"Z{row_num}/AA{row_num}={today}."
                )
        else:
            # NOT FOUND: add new OUT row with 50xxx sequence
            row_num = ws.max_row + 1
            seq = next_new_seq
            next_new_seq += 1
            new_items_count += 1

            copy_cell_styles(ws, STYLE_TEMPLATE_ROW, row_num, ["A", "C", "I", "X", "Z", "AA"])

            ws[f"A{row_num}"].value = seq      # sequential 50xxx+
            ws[f"C{row_num}"].value = code     # Asset Id
            ws[f"I{row_num}"].value = loc
            ws[f"X{row_num}"].value = "Out"
            ws[f"Z{row_num}"].value = today
            ws[f"AA{row_num}"].value = today

            normalized_asset_set.add(normalized_code)
            normalized_to_row[normalized_code] = row_num
            scanned.add(normalized_code)

            log(
                f"[X] {code} NOT found in inventory. "
                f"Added OUT row {row_num} with A={seq}, C={code}, I='{loc}', "
                f"X=Out, Z/AA={today}."
            )

    ttk.Button(btns, text="Finish (Done)", command=save_and_close).pack(side="left")
    ttk.Button(
        btns,
        text="Clear Entry",
        command=lambda: scan_entry.delete(0, tk.END),
    ).pack(side="left", padx=8)
    ttk.Button(btns, text="Quit All", command=close_window).pack(side="right")

    scan_entry.bind("<Return>", handle_scan)
    scan_win.bind("<Escape>", lambda e: close_window())

    # Start loop in standalone mode
    if owns_root:
        root.deiconify()
        root.mainloop()


if __name__ == "__main__":
    run_scan_out()